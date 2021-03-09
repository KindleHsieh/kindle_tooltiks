import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from copy import copy
import numpy as np


# Read solider data.
solider_data = []
wb = xl.load_workbook("役男七月餐假.xlsx")
sh = wb.sheetnames
sh0 = sh[0]
print(sh0)
rows = wb[sh0].rows
for i in rows:
    dd = []
    if i[0].value is not None:
        for j in i:
            #print(j.value, end=",")
            dd += [j.value]
        #print()
        solider_data.append(dd)
del wb, sh, sh0, rows, dd

# Input
date = ['108', 7, 1]  # Year & Month & first day week. (Y/M/W)
workday = []  # make-up work
holiday = []  # national holiday
com_leave = []  # compatible leave

# Count how many days in a month.
if date[1] == 1 or date[1] == 3 or date[1] == 5 or date[1] == 7 or date[1] == 8 or date[1] == 10 or date[1] == 12:
    days = 31
elif date[1] == 2:
    days = 28
else:
    days = 30
what_day = ("一", "二", "三", "四", "五", "六", "日")

# Setting cell styles.  (Material)
font1 = Font(bold=True, size=16, name='標楷體')
font2 = Font(name='標楷體', size=12, color=xl.styles.colors.RED)
font3 = Font(name='標楷體', size=14)
font4 = Font(name='標楷體', size=12)
font5 = Font(name='標楷體', size=20)
font6 = Font(name='標楷體', size=14, bold=True)

fill1 = PatternFill('solid', fgColor='00E3AF86')
border1 = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'),
                 bottom=Side(border_style='thin'))
border2 = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'),
                 bottom=Side(border_style='double'))
# bottom_left.
border3 = Border(left=Side(border_style='medium'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='medium'))
# top.
border4 = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='medium'), bottom=Side(border_style='thin'))
# bottom_right.
border5 = Border(left=Side(border_style='thin'), right=Side(border_style='medium'), top=Side(border_style='thin'), bottom=Side(border_style='medium'))
# bottom.
border6 = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='medium'))
# top_right.
border7 = Border(left=Side(border_style='thin'), right=Side(border_style='medium'), top=Side(border_style='medium'), bottom=Side(border_style='thin'))
# top_left.
border8 = Border(left=Side(border_style='medium'), right=Side(border_style='thin'), top=Side(border_style='medium'), bottom=Side(border_style='thin'))
# right.
border9 = Border(left=Side(border_style='thin'), right=Side(border_style='medium'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))

align1 = Alignment(horizontal='center', vertical='center')
align2 = Alignment(horizontal='center', vertical='center', wrap_text=True)

## Create an Excel file.
wb = Workbook()
###################################################################################################
#  Create a meals sheet.
ws_meal = wb.create_sheet('meal', 0)

# Setting cell styles.  (Merge\Cell styles)
ws_meal.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days+3)
ws_meal.merge_cells(start_row=4+3*(len(solider_data)-1), start_column=1, end_row=4+3*(len(solider_data)-1)+3, end_column=1)
for i in range(4, 3*len(solider_data), 3):
    ws_meal.merge_cells(start_row=i, start_column=1, end_row=i+2, end_column=1)

for i in range(1, 3*len(solider_data)+4+1):
    for j in range(1, days+2+1+1):
        ws_meal.cell(i, j).border = border1
        ws_meal.cell(i, j).alignment = align1
        ws_meal.cell(i, j).font = font4
for i in range(4, 3*len(solider_data), 3):
    for j in range(1, days+3+1):
        ws_meal.cell(i + 2, j).border = border2

ws_meal.cell(1, 1).value = "衛生福利部南區老人之家替代役役男"+date[0]+"年"+str(date[1])+"月份餐費明細"
ws_meal.cell(1, 1).font = font1
# which years and months.
ws_meal.cell(2, 1).value = str(int(date[0])+1911)+"年"
ws_meal.cell(2, 1).font = font3
ws_meal.cell(2, 2).value = str(date[1])+"月"
ws_meal.cell(2, 2).font = font3
# Write the days and week.
index = date[2]-1
day_list = []
week_list = []
for i in range(days):
    ws_meal.cell(2, i+3).value = i+1
    ws_meal.cell(3, i+3).value = what_day[index]
    ws_meal.cell(2, i + 3).font = font3
    ws_meal.cell(3, i + 3).font = font4
    day_list += [i + 1]
    week_list += [index+1]
    index += 1
    if index == 7:
        index = 0
print(day_list)
print(week_list)
# Write the common words.
ws_meal.cell(3, days+3).value = "小計"
for i in range(4,3*(len(solider_data)+1), 3):
    ws_meal.cell(i,2).value = '早餐'
    ws_meal.cell(i+1, 2).value = '中餐'
    ws_meal.cell(i+2, 2).value = '晚餐'
ws_meal.cell(4+3*(len(solider_data)-1), 1).value = '總計'
ws_meal.cell(4+3*(len(solider_data)-1)+3, 2).value = '總和'

# Write down the meals.
# Mark up the red day on the calendar.
work_end = []
free_day = holiday + com_leave
free_day.extend([day_list[e] for e,i in enumerate(week_list) if (i == 6 or i == 7)])  # Sat. an Sun.
free_day = [i for i in free_day if i not in workday]
print(free_day)
for i in free_day:
    if (i-1) not in free_day and (i-1) != 0:
        work_end.append(i-1)
# If the end of this month is Friday, appending this day into work_end.
work_end.append(day_list[-1]) if week_list[-1] == 5 else print(day_list[-1])
for i in free_day:
    ws_meal.cell(3,3+i-1).font = font2
    for j in range(3, 3*len(solider_data)+4+1):
        ws_meal.cell(j, 3+i-1).fill = fill1


for i in range(1, len(solider_data)):
    breakfast = 0
    lunch = 0
    dinner = 0
    month_free = []
    com_free = []
    Free = []
    extra = []
    logout_day = []
    # Log out date.
    logout_day = solider_data[i][7].date().strftime('%Y-%m-%d')
    if int(logout_day[:4]) == int(date[0])+1911 and int(logout_day[5:7]) == date[1]:
        logout_day = int(logout_day[-2:]) - 1
    else:
        logout_day = days
    # Solider name.
    ws_meal.cell(4+3*(i-1), 1).value = solider_data[i][0]
    ws_meal.cell(4+3*(i-1), 1).font = font3
    # Normal meals.
    if solider_data[i][1] == 'Y' and solider_data[i][2] == 'Y':
        for j in range(1, logout_day + 1):
            if j not in free_day:
                breakfast += 1
                lunch += 1
                dinner += 1
                ws_meal.cell(4 + 3 * (i - 1), 3 + j - 1).value = 1
                ws_meal.cell(4 + 3 * (i - 1) + 1, 3 + j - 1).value = 1
                ws_meal.cell(4 + 3 * (i - 1) + 2, 3 + j - 1).value = 1
    elif solider_data[i][1] == 'N' and solider_data[i][2] == 'Y':
        for j in range(1, logout_day + 1):
            if j not in free_day:
                lunch += 1
                dinner += 1
                ws_meal.cell(4 + 3 * (i - 1) + 1, 3 + j - 1).value = 1
                ws_meal.cell(4 + 3 * (i - 1) + 2, 3 + j - 1).value = 1
    elif solider_data[i][1] == 'Y' and solider_data[i][2] == 'N':
        for j in range(1, logout_day + 1):
            if j not in free_day:
                breakfast += 1
                ws_meal.cell(4 + 3 * (i - 1), 3 + j - 1).value = 1
    # Work end meal.
    if solider_data[i][3] == 'N':
        for j in work_end:
            if j <= logout_day:
                ws_meal.cell(4 + 3 * (i - 1) + 2, 3 + j - 1).value = None
                dinner -= 1
    # Solider free day.
    if solider_data[i][4] is not None:
        if type(solider_data[i][4]) is str:
            month_free = list(map(int, solider_data[i][4].split(',')))
        elif type(solider_data[i][4]) is int:
            month_free = [solider_data[i][4]]
    if solider_data[i][5] is not None:
        if type(solider_data[i][5]) is str:
            com_free = list(map(int, solider_data[i][5].split(',')))
        elif type(solider_data[i][5]) is int:
            com_free = [solider_data[i][5]]
    Free += month_free
    Free += com_free

    for j in Free:
        if j in work_end:
            ws_meal.cell(4 + 3 * (i - 1), 3 + j - 1).value = None
            ws_meal.cell(4 + 3 * (i - 1) + 1, 3 + j - 1).value = None
            breakfast -= 1
            lunch -= 1
        else:
            ws_meal.cell(4 + 3 * (i - 1),     3 + j - 1).value = None
            ws_meal.cell(4 + 3 * (i - 1) + 1, 3 + j - 1).value = None
            ws_meal.cell(4 + 3 * (i - 1) + 2, 3 + j - 1).value = None
            breakfast -= 1
            lunch -= 1
            dinner -= 1
    # Extra meals.
    if solider_data[i][6] is not None:
        if type(solider_data[i][6]) is str:
            extra = list(map(int,solider_data[i][6].split(',')))
        elif type(solider_data[i][6]) is int:
            extra = [solider_data[i][6]]
        # How to add the meals.
        if solider_data[i][1] == 'Y' and solider_data[i][2] == 'Y':
            for j in extra:
                breakfast += 1
                lunch += 1
                dinner += 1
                ws_meal.cell(4 + 3 * (i - 1), 3 + j - 1).value = 1
                ws_meal.cell(4 + 3 * (i - 1) + 1, 3 + j - 1).value = 1
                ws_meal.cell(4 + 3 * (i - 1) + 2, 3 + j - 1).value = 1
        elif solider_data[i][1] == 'N' and solider_data[i][2] == 'Y':
            for j in extra:
                lunch += 1
                dinner += 1
                ws_meal.cell(4 + 3 * (i - 1) + 1, 3 + j - 1).value = 1
                ws_meal.cell(4 + 3 * (i - 1) + 2, 3 + j - 1).value = 1
        elif solider_data[i][1] == 'Y' and solider_data[i][2] == 'N':
            for j in extra:
                breakfast += 1
                ws_meal.cell(4 + 3 * (i - 1), 3 + j - 1).value = 1
    # count of meals.
    breakfast = 0 if breakfast < 0 else breakfast
    lunch = 0 if lunch < 0 else lunch
    dinner = 0 if dinner < 0 else dinner
    ws_meal.cell(4 + 3 * (i - 1), 3 + days).value = breakfast
    ws_meal.cell(4 + 3 * (i - 1) + 1, 3 + days).value = lunch
    ws_meal.cell(4 + 3 * (i - 1) + 2, 3 + days).value = dinner

columns = ws_meal.columns


for i, e in enumerate(columns):
    sum_breakfast = 0
    sum_lunch = 0
    sum_dinner = 0
    if 1 < i < 1+days+2:
        sum_breakfast = sum([k.value for j, k in enumerate(e) if k.value is not None if j in range(3, 3 + 3 * len(solider_data) + 1, 3)])
        ws_meal.cell(4 + 3 * (len(solider_data) - 1), i + 1).value = sum_breakfast
        sum_lunch = sum([k.value for j, k in enumerate(e) if k.value is not None if j in range(4, 4 + 3 * len(solider_data) + 2, 3)])
        ws_meal.cell(5 + 3 * (len(solider_data) - 1), i + 1).value = sum_lunch
        sum_dinner = sum([k.value for j, k in enumerate(e) if k.value is not None if j in range(5, 5 + 3 * len(solider_data) + 3, 3)])
        ws_meal.cell(6 + 3 * (len(solider_data) - 1), i + 1).value = sum_dinner
        ws_meal.cell(7 + 3 * (len(solider_data) - 1), i + 1).value = sum_breakfast + sum_lunch + sum_dinner

#########################################################################################################
## Create a kitchen sheet.
ws_kitchen = wb.create_sheet('kitchen', 1)
ws_kitchen.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3+days)
for i in range(len(solider_data)):
    ws_kitchen.merge_cells(start_row=3+i, start_column=1, end_row=3+i, end_column=2)
ws_kitchen.merge_cells(start_row=3+len(solider_data), start_column=1, end_row=4+len(solider_data), end_column=1)
for i in range(1, 3+len(solider_data)+2):
    for j in range(1, days+2+1+1):
        ws_kitchen.cell(i, j).border = border1
        ws_kitchen.cell(i, j).alignment = align1
        ws_kitchen.cell(i, j).font = font5 # Size=20

for i in range(1, 4):
    for j in range(1, 4+days):
        ws_kitchen.cell(i, j).value = ws_meal.cell(i, j).value
        ws_kitchen.cell(i, j).border = copy(ws_meal.cell(i, j).border)
        ws_kitchen.cell(i, j).alignment = copy(ws_meal.cell(i, j).alignment)
ws_kitchen.cell(1, 1).font = Font(name='標楷體', size=20, bold=True)
for i in free_day:
    ws_kitchen.cell(3, 3+i-1).font = Font(name='標楷體', size=20, color=xl.styles.colors.RED)
    for j in range(3, 3+len(solider_data)+2):
        ws_kitchen.cell(j, 3+i-1).fill = fill1


for i, e in enumerate(range(4, 4+3*(len(solider_data)-1)+1, 3), 4):
    ws_kitchen.cell(i, 1).value = ws_meal.cell(e, 1).value

for i, e in enumerate(range(5, 5+3*(len(solider_data)-1)+1, 3), 4):
    for j in range(3, 3+days+1):
        ws_kitchen.cell(i, j).value = ws_meal.cell(e, j).value

for i in work_end:
    ws_kitchen.cell(2+len(solider_data)+2, 2+i).value = ws_meal.cell(4+3*(len(solider_data)-1)+2, 2+i).value

#########################################################################################################
# Meal fee
ws_fee = wb.create_sheet('fee', 2)
ws_fee.merge_cells('A1:H1')
for i in range(3, 3*(len(solider_data)-1)+1, 3):
    for j in [1, 2, 7]:
        ws_fee.merge_cells(start_row=i, start_column=j, end_row=i+2, end_column=j)

ws_fee.merge_cells(start_row=3+3*(len(solider_data)-1), start_column=1, end_row=3+3*(len(solider_data)-1), end_column=6)
ws_fee.merge_cells(start_row=3+3*(len(solider_data)-1)+1, start_column=7, end_row=3+3*(len(solider_data)-1)+1, end_column=8)
# all
for i in range(1, 3*(len(solider_data))+1):
    for j in range(1, 9):
        ws_fee.cell(i, j).border = border1
        ws_fee.cell(i, j).font = font4
# left_bottom.
for i in range(2, 3*len(solider_data)+1):
    ws_fee.cell(i, 1).border = border3
# top.
for i in range(2, 8):
    ws_fee.cell(1, i).border = border4
# right_bottom.
for i in range(2, 3*len(solider_data)+1, 3):
    ws_fee.cell(i, 8).border = border5
ws_fee.cell(3*len(solider_data), 8).border = border5
# right.
for i in range(3, 3*(len(solider_data)-1)+1, 3):
    ws_fee.cell(i, 8).border = border9
    ws_fee.cell(i+1, 8).border = border9
# bottom.
for i in range(2, 3*len(solider_data)+1, 3):
    for j in range(2, 8):
        ws_fee.cell(i, j).border = border6
for i in range(2, 8):
    ws_fee.cell(3*len(solider_data), i).border = border6
# top_right.
ws_fee.cell(1, 8).border = border7
# top_left.
ws_fee.cell(1, 1).border = border8


# normal word.
ws_fee.cell(1, 1).value = '衛生福利部南區老人之家替代役役男108年' + str(date[1]) + '月份搭伙明細'
ws_fee.column_dimensions['H'].width = 40
ws_fee.cell(1, 1).font = font1
ws_fee.cell(1, 1).alignment = align1

ws_fee.cell(2, 1).value = '編號'
ws_fee.cell(2, 2).value = '姓名'
ws_fee.cell(2, 3).value = '時段'
ws_fee.cell(2, 4).value = '日數'
ws_fee.cell(2, 5).value = '金額/餐'
ws_fee.cell(2, 6).value = '餐總金額'
ws_fee.cell(2, 7).value = '總金額'

for i in range(1, 8):
    ws_fee.cell(2, i).alignment = align1

for i in range(1, len(solider_data)):
    ws_fee.cell(3*i, 1).value = i
    ws_fee.cell(3*i, 2).value = solider_data[i][0]
    ws_fee.cell(3*i, 1).alignment = align1
    ws_fee.cell(3*i, 2).alignment = align1
    ws_fee.cell(3*i, 7).alignment = align1

    ws_fee.cell(3*i, 3).value = '早餐'
    ws_fee.cell(3*i+1, 3).value = '中餐'
    ws_fee.cell(3*i+2, 3).value = '晚餐'

    ws_fee.cell(3 * i, 4).value = ws_meal.cell(3 * i + 1, 3 + days).value
    ws_fee.cell(3 * i+1, 4).value = ws_meal.cell(3 * i + 2, 3 + days).value
    ws_fee.cell(3 * i+2, 4).value = ws_meal.cell(3 * i + 3, 3 + days).value

    ws_fee.cell(3*i, 5).value = 45
    ws_fee.cell(3*i+1, 5).value = 45
    ws_fee.cell(3*i+2, 5).value = 50

    ws_fee.cell(3 * i, 6).value = ws_fee.cell(3 * i, 4).value * ws_fee.cell(3*i, 5).value
    ws_fee.cell(3 * i + 1, 6).value = ws_fee.cell(3 * i+1, 4).value * ws_fee.cell(3*i+1, 5).value
    ws_fee.cell(3 * i + 2, 6).value = ws_fee.cell(3 * i+2, 4).value * ws_fee.cell(3*i+2, 5).value

    ws_fee.cell(3 * i, 7).value = ws_fee.cell(3 * i, 6).value + ws_fee.cell(3 * i+1, 6).value + ws_fee.cell(3 * i+2, 6).value

ws_fee.cell(3*len(solider_data), 7).value = sum([ws_fee.cell(i, 7).value for i in range(3, 3*(len(solider_data)-1)+1, 3)])
ws_fee.cell(3*len(solider_data), 7).alignment = align1

ws_fee.cell(3*len(solider_data), 1).value = '全體役男月餐費總金額'
ws_fee.cell(3*len(solider_data), 1).alignment = align1
ws_fee.cell(3*len(solider_data)+1, 1).value = '製表'
ws_fee.cell(3*len(solider_data)+1, 3).value = '承辦人'
ws_fee.cell(3*len(solider_data)+1, 5).value = '少教科'
ws_fee.cell(3*len(solider_data)+1, 7).value = '出納               主計機構                主任'

#########################################################################################################
# Shift table.
ws_shift = wb.create_sheet('shift table', 3)
ws_shift.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2+len(solider_data))
ws_shift.merge_cells(start_row=3, start_column=2+len(solider_data), end_row=3+days-1+3, end_column=2+len(solider_data))
ws_shift.merge_cells(start_row=2+days+1, start_column=1, end_row=2+days+1, end_column=2)
ws_shift.merge_cells(start_row=2+days+2, start_column=1, end_row=2+days+2, end_column=2)
ws_shift.merge_cells(start_row=2+days+3, start_column=1, end_row=2+days+3, end_column=2)
# Border.
for i in range(1, 2+days+3+1):
    for j in range(1, 2+len(solider_data)+1):
        ws_shift.cell(i, j).border = border1
        ws_shift.cell(i, j).alignment = align1
# Font.
for i in range(1, 2+len(solider_data)+1):
    ws_shift.cell(2, i).font = font6
    ws_shift.cell(2, i).alignment = align2

for i in range(3,3+days+3):
    ws_shift.cell(i, 1).font = font4
    ws_shift.cell(i, 2).font = font4

for i in range(1, 2+len(solider_data)+1):
    ws_shift.cell(2+days+3, i).font = Font(name='標楷體', size=12, bold=True)

for i in range(3, 3+days):
    for j in range(3, 3+len(solider_data)):
        ws_shift.cell(i, j).font = Font(name='標楷體', size=12, bold=True)
# Normal.
ws_shift.cell(1, 1).value = "衛生福利部南區老人之家" + date[0] + "年" + str(date[1]) + "月替代役早班輪值表"
ws_shift.cell(1, 1).font = Font(name='標楷體', size=18, bold=True)
ws_shift.cell(1, 1).alignment = align1

ws_shift.cell(2, 1).value = '值\n勤\n日\n期'
ws_shift.cell(2, 1).font = font3

ws_shift.cell(2, 2).value = '姓\n名'
ws_shift.cell(2, 2).font = font3

ws_shift.cell(2, 2+len(solider_data)).value = '備考'
ws_shift.cell(2, 2+len(solider_data)).font = font3

ws_shift.cell(2+days+1, 1).value = '本月可休假天數'
ws_shift.cell(2+days+2, 1).value = '本月實際休假天數'
ws_shift.cell(2+days+3, 1).value = '本月底積假天數'

ws_shift.cell(3, 2+len(solider_data)).value = '1."甲"為早班，上班時間0730至1130與1330至1730。\n2."丁"為休假。\n需按本表值勤，若有必要需調班時，請事先告知管理幹部。'
ws_shift.cell(3, 2+len(solider_data)).alignment = align2
ws_shift.cell(3, 2+len(solider_data)).font = Font(name='標楷體', size=12)
# Name.
for i in range(3, 3+len(solider_data)-1):
    strr=''
    for j in range(len(solider_data[i-2][0])):
        strr += solider_data[i-2][0][j] + '\n'
    ws_shift.cell(2, i).value = strr[:-1]
del strr
# 甲 or 丁.
for i in range(1, len(solider_data)):
    on_duty = []
    logout_day = []
    month_free = []
    free = 0
    if solider_data[i][8] is not None:
        if type(solider_data[i][8]) is str:
            on_duty = list(map(int, solider_data[i][8].split(',')))
        elif type(solider_data[i][8]) is int:
            on_duty = [solider_data[i][8]]

    logout_day = solider_data[i][7].date().strftime('%Y-%m-%d')
    if int(logout_day[:4]) == int(date[0]) + 1911 and int(logout_day[5:7]) == date[1]:
        logout_day = int(logout_day[-2:]) - 1
    else:
        logout_day = days

    for j in range(3, 3+logout_day):
        if j-2 in free_day:
            if j-2 in on_duty:
                ws_shift.cell(j, i + 2).value = '甲'
            else:
                ws_shift.cell(j, i + 2).value = '丁'
                ws_shift.cell(j, i + 2).font = Font(name='標楷體', size=12, color=xl.styles.colors.RED, bold=True)
        else:
            ws_shift.cell(j, i+2).value = '甲'

    if solider_data[i][4] is not None:
        if type(solider_data[i][4]) is str:
            month_free = list(map(int, solider_data[i][4].split(',')))
        elif type(solider_data[i][4]) is int:
            month_free = [solider_data[i][4]]

    for j in month_free:
        ws_shift.cell(j+2, i+2).value = '丁'
        ws_shift.cell(j+2, i + 2).font = Font(name='標楷體', size=12, color=xl.styles.colors.RED, bold=True)
# How many days.
    for j in range(3, 3+days):
        if ws_shift.cell(j, 2+i).value == '丁':
            free += 1
    print(free)
    print(free_day)
    print(logout_day)
    ws_shift.cell(2+days+1, 2+i).value = len([k for k in free_day if k <= logout_day])
    ws_shift.cell(2+days+2, 2+i).value = free
    ws_shift.cell(2+days+3, 2+i).value = len([k for k in free_day if k <= logout_day]) - free

# Date and Week.
for i in range(3, 3+days):
    ws_shift.cell(i, 1).value = ws_meal.cell(2, i).value
    ws_shift.cell(i, 2).value = ws_meal.cell(3, i).value

ws_shift.freeze_panes = ws_shift['C3']
# Save the file.
wb.save("2019_7v1.xlsx")